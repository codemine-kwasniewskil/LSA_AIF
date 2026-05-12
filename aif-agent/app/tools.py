"""
Tool implementations for the AIF Knowledge Agent.

Five tools are exposed to Claude:
  1. search_interfaces      — find interfaces by keyword in xlsx
  2. get_interface_details  — full metadata from all 4 xlsx tables
  3. get_interface_readme   — pre-generated README documentation
  4. search_abap_code       — grep across all .abap files
  5. read_abap_file         — return full source of one .abap file

All file I/O is synchronous. Call from async code via asyncio.to_thread().
"""

import os
import re
from pathlib import Path

import openpyxl

# ── paths ─────────────────────────────────────────────────────────────────────

DATA_DIR = Path(os.environ.get("DATA_DIR", "data"))
XLSX_DIR = DATA_DIR / "xlsx"
DOCS_DIR = DATA_DIR / "docs"
ABAP_DIR = DATA_DIR / "abapgit"

# ── xlsx helpers ──────────────────────────────────────────────────────────────


def _find_columns(ws) -> dict[str, int]:
    """Return {HEADER_NAME_UPPER: 0-based-column-index} from the first row."""
    cols: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for i, cell in enumerate(row):
            if cell is not None:
                cols[str(cell).strip().upper()] = i
        break
    return cols


def _v(row: tuple, idx: int | None, default: str = "") -> str:
    """Safely extract a cell value as a stripped string."""
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    return str(val).strip() if val is not None else default


def _load_wb(filename: str):
    """Load a workbook from XLSX_DIR. Returns None if file not found."""
    path = XLSX_DIR / filename
    if not path.exists():
        return None
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _missing(filename: str) -> str:
    return (
        f"Data file '{filename}' not found in {XLSX_DIR}. "
        f"Run setup_data.py to copy data files."
    )


# ── Tool 1: search_interfaces ─────────────────────────────────────────────────


def tool_search_interfaces(query: str) -> str:
    """Search all interface definitions for a keyword."""
    wb = _load_wb("aif_t_finf_de.xlsx")
    if wb is None:
        return _missing("aif_t_finf_de.xlsx")

    ws = wb.active
    cols = _find_columns(ws)

    ns_i = cols.get("NS")
    ifname_i = cols.get("IFNAME")
    ver_i = cols.get("IFVERSION")
    desc_i = cols.get("IFDESC")

    q = query.lower()
    results: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ns = _v(row, ns_i)
        ifname = _v(row, ifname_i)
        ver = _v(row, ver_i)
        desc = _v(row, desc_i)

        if not ns and not ifname:
            continue

        if q in ns.lower() or q in ifname.lower() or q in desc.lower():
            results.append(f"NS={ns} | IFNAME={ifname} | VERSION={ver} | {desc}")

    wb.close()

    if not results:
        return f"No interfaces found matching '{query}'."

    header = f"Found {len(results)} interface(s) matching '{query}':\n"
    return header + "\n".join(results[:50])


# ── Tool 2: get_interface_details ─────────────────────────────────────────────


def tool_get_interface_details(ns: str, ifname: str) -> str:
    """Return full metadata for one interface from all 4 AIF config tables."""
    ns_u = ns.strip().upper()
    ifname_u = ifname.strip().upper()
    out: list[str] = []

    # ── 1. Interface definition (aif_t_finf_de) ──────────────────────────────
    wb = _load_wb("aif_t_finf_de.xlsx")
    if wb is None:
        out.append(_missing("aif_t_finf_de.xlsx"))
    else:
        ws = wb.active
        cols = _find_columns(ws)
        ns_i = cols.get("NS")
        ifname_i = cols.get("IFNAME")
        ver_i = cols.get("IFVERSION")
        desc_i = cols.get("IFDESC")
        ddic_i = cols.get("DDICSTRUCTURE") or cols.get("DDIC_STRUCTURE")
        check_i = cols.get("FUBA_CHECK")
        init_i = cols.get("FUBA_INIT")

        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _v(row, ns_i) == ns_u and _v(row, ifname_i) == ifname_u:
                out.append("## 1. Interface Definition (AIF_T_FINF)")
                out.append(f"NS        : {_v(row, ns_i)}")
                out.append(f"IFNAME    : {_v(row, ifname_i)}")
                out.append(f"VERSION   : {_v(row, ver_i)}")
                out.append(f"DESC      : {_v(row, desc_i)}")
                if ddic_i is not None:
                    out.append(f"DDICSTRUCT: {_v(row, ddic_i)}")
                if check_i is not None:
                    out.append(f"FUBA_CHECK: {_v(row, check_i)}")
                if init_i is not None:
                    out.append(f"FUBA_INIT : {_v(row, init_i)}")
                found = True
                break

        if not found:
            out.append(f"## 1. Interface Definition\nNot found: NS={ns_u} IFNAME={ifname_u}")
        wb.close()

    out.append("")

    # ── 2. Actions (aif_t_ifact) ──────────────────────────────────────────────
    wb = _load_wb("aif_t_ifact.xlsx")
    if wb is None:
        out.append(_missing("aif_t_ifact.xlsx"))
    else:
        ws = wb.active
        cols = _find_columns(ws)
        ns_i = cols.get("NS")
        ifname_i = cols.get("IFNAME")
        act_i = cols.get("ACTIONNR")
        nsact_i = cols.get("NSACTION")
        ifact_i = cols.get("IFACTION")
        stop_i = cols.get("STOP_ON_ERROR") or cols.get("STOPONERROR")

        rows_found: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _v(row, ns_i) == ns_u and _v(row, ifname_i) == ifname_u:
                actionnr = _v(row, act_i)
                nsaction = _v(row, nsact_i)
                ifaction = _v(row, ifact_i)
                stop = _v(row, stop_i)
                rows_found.append(
                    f"  ACTIONNR={actionnr:>4} | NS={nsaction} | ACTION={ifaction:<30} | STOP_ON_ERROR={stop}"
                )

        out.append("## 2. Action Pipeline (AIF_T_IFACT)")
        if rows_found:
            out.extend(sorted(rows_found))
        else:
            out.append("  No actions found.")
        wb.close()

    out.append("")

    # ── 3. Function modules per action (aif_t_func) ───────────────────────────
    # AIF_T_FUNC has no IFNAME column — it is keyed on NS + IFACTION (action level).
    # We first collect (NSACTION, IFACTION) pairs from the AIF_T_IFACT rows already read,
    # then look up AIF_T_FUNC by those pairs.
    #
    # Actual column name for sequence is FUNCNR (not SEQNR).

    # Collect action pairs from the ifact workbook (re-open for a second pass)
    action_pairs: set[tuple[str, str]] = set()
    wb_act2 = _load_wb("aif_t_ifact.xlsx")
    if wb_act2 is not None:
        ws_act2 = wb_act2.active
        cols_act2 = _find_columns(ws_act2)
        ns_i2 = cols_act2.get("NS")
        ifname_i2 = cols_act2.get("IFNAME")
        nsact_i2 = cols_act2.get("NSACTION")
        ifact_i2 = cols_act2.get("IFACTION")
        for row in ws_act2.iter_rows(min_row=2, values_only=True):
            if _v(row, ns_i2) == ns_u and _v(row, ifname_i2) == ifname_u:
                action_pairs.add((_v(row, nsact_i2), _v(row, ifact_i2)))
        wb_act2.close()

    wb = _load_wb("aif_t_func.xlsx")
    if wb is None:
        out.append(_missing("aif_t_func.xlsx"))
    else:
        ws = wb.active
        cols = _find_columns(ws)
        ns_i = cols.get("NS")
        ifact_i = cols.get("IFACTION")
        func_i = cols.get("FUNCTION")
        seq_i = cols.get("FUNCNR")          # correct column name in this table
        stop_i = cols.get("STOP_ON_ERROR")

        rows_found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_ns = _v(row, ns_i)
            row_act = _v(row, ifact_i)
            if (row_ns, row_act) in action_pairs:
                funcnr = _v(row, seq_i)
                func = _v(row, func_i)
                stop = _v(row, stop_i)
                rows_found.append(
                    f"  FUNCNR={funcnr:>4} | NS={row_ns} | ACTION={row_act:<30} | FM={func} | STOP={stop}"
                )

        out.append("## 3. Function Modules (AIF_T_FUNC)")
        if rows_found:
            out.extend(sorted(rows_found))
        else:
            out.append("  No function modules found.")
        wb.close()

    out.append("")

    # ── 4. Field mappings (aif_t_fmap) ────────────────────────────────────────
    wb = _load_wb("aif_t_fmap.xlsx")
    if wb is None:
        out.append(_missing("aif_t_fmap.xlsx"))
    else:
        ws = wb.active
        cols = _find_columns(ws)
        ns_i = cols.get("NS")
        ifname_i = cols.get("IFNAME")
        smapnr_i = cols.get("SMAPNR")
        vmap_i = cols.get("VMAPNAME")
        check_i = cols.get("AIFCHECK")
        # Actual column names in aif_t_fmap: FIELDNAME (source) and SAP_FIELDNAME1 (target)
        src_i = cols.get("FIELDNAME")
        tgt_i = cols.get("SAP_FIELDNAME1")

        rows_found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _v(row, ns_i) == ns_u and _v(row, ifname_i) == ifname_u:
                smapnr = _v(row, smapnr_i)
                vmap = _v(row, vmap_i)
                check = _v(row, check_i)
                src = _v(row, src_i)
                tgt = _v(row, tgt_i)
                rows_found.append(
                    f"  SMAPNR={smapnr:<4} | VMAP={vmap:<30} | CHECK={check:<10} | FIELDNAME={src} -> SAP_FIELD={tgt}"
                )

        out.append("## 4. Field Mappings (AIF_T_FMAP)")
        if rows_found:
            out.extend(rows_found[:80])  # cap at 80 rows
            if len(rows_found) > 80:
                out.append(f"  ... ({len(rows_found) - 80} more rows truncated)")
        else:
            out.append("  No field mappings found.")
        wb.close()

    return "\n".join(out)


# ── Tool 3: get_interface_readme ──────────────────────────────────────────────


def tool_get_interface_readme(name: str) -> str:
    """Return pre-generated README markdown for an interface or domain."""
    if not DOCS_DIR.exists():
        return f"Docs directory not found: {DOCS_DIR}. Run setup_data.py first."

    name_clean = name.strip().upper()

    # Candidate filenames to try (most specific first)
    candidates = [
        f"{name_clean}_Interface_README.md",
        f"{name_clean}_README.md",
        f"{name_clean}.md",
    ]

    # Also search by partial match across all md files
    all_md = list(DOCS_DIR.glob("*.md"))

    # Exact candidates first
    for cand in candidates:
        path = DOCS_DIR / cand
        if path.exists():
            content = path.read_text(encoding="utf-8", errors="replace")
            return f"File: {path.name}\n\n{content}"

    # Partial match — filename contains the search term
    for md_path in sorted(all_md):
        if name_clean in md_path.name.upper():
            content = md_path.read_text(encoding="utf-8", errors="replace")
            return f"File: {md_path.name}\n\n{content}"

    available = ", ".join(p.name for p in sorted(all_md))
    return (
        f"No README found matching '{name}'.\n"
        f"Available docs: {available}\n"
        f"Use get_interface_details to retrieve data directly from xlsx tables."
    )


# ── Tool 4: search_abap_code ──────────────────────────────────────────────────

# All text-based abapgit file extensions to include in searches
_ABAP_EXTENSIONS = (
    "*.abap",           # function modules, programs, includes, classes
    "*.asddls",         # CDS views
    "*.xslt.source.xml",# XSLT transformations
    "*.fugr.xml",       # function group metadata
    "*.tabl.xml",       # DDIC table/structure definitions
    "*.ddls.xml",       # CDS view metadata
    "*.dtel.xml",       # data element definitions
    "*.doma.xml",       # domain definitions
    "*.tran.xml",       # transaction definitions
    "*.clas.xml",       # class metadata
    "*.prog.xml",       # program metadata
)


def _iter_all_abap_files(base: Path):
    """Yield all text-based abapgit files under base, sorted by path."""
    seen: set[Path] = set()
    for pattern in _ABAP_EXTENSIONS:
        for p in base.rglob(pattern):
            if p not in seen:
                seen.add(p)
                yield p


def tool_search_abap_code(query: str, context_lines: int = 15) -> str:
    """Full-text search across all abapgit source files. Returns matching lines with context."""
    if not ABAP_DIR.exists():
        return f"abapgit directory not found: {ABAP_DIR}. Run setup_data.py first."

    all_files = sorted(_iter_all_abap_files(ABAP_DIR), key=lambda p: str(p))
    if not all_files:
        return f"No abapgit source files found under {ABAP_DIR}."

    query_lower = query.lower()
    output_parts: list[str] = []
    total_chars = 0
    MAX_CHARS = 12_000

    for abap_file in all_files:
        try:
            content = abap_file.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue

        lines = content.splitlines()
        match_positions = [
            i for i, line in enumerate(lines) if query_lower in line.lower()
        ]

        if not match_positions:
            continue

        rel_path = abap_file.relative_to(ABAP_DIR)
        file_block = [f"\n### {rel_path}"]

        # Show up to 3 matches per file with context
        for pos in match_positions[:3]:
            start = max(0, pos - context_lines)
            end = min(len(lines), pos + context_lines + 1)
            file_block.append(f"  (match at line {pos + 1})")
            for j, line_text in enumerate(lines[start:end], start=start + 1):
                marker = ">>>" if j == pos + 1 else "   "
                file_block.append(f"  {marker} {j:4d} | {line_text}")

        block_text = "\n".join(file_block)
        total_chars += len(block_text)
        output_parts.append(block_text)

        if total_chars >= MAX_CHARS:
            output_parts.append("\n... (output truncated — refine your search term)")
            break

    if not output_parts:
        return f"No matches found for '{query}' in {len(all_files)} abapgit files."

    header = f"Search: '{query}' — found in {len(output_parts)} file(s)\n"
    return header + "\n".join(output_parts)


# ── Tool 5: read_abap_file ────────────────────────────────────────────────────


def tool_read_abap_file(filename: str) -> str:
    """Read the full source of any abapgit file matched by partial name."""
    if not ABAP_DIR.exists():
        return f"abapgit directory not found: {ABAP_DIR}. Run setup_data.py first."

    name_lower = filename.strip().lower()
    # Normalise namespace separator so user can type either / or #
    name_lower = name_lower.replace("/thkr/", "#thkr#").replace("/", "#")

    matches: list[Path] = []
    for abap_file in _iter_all_abap_files(ABAP_DIR):
        if name_lower in abap_file.name.lower():
            matches.append(abap_file)

    if not matches:
        # Try a more lenient search — strip namespace prefix from query
        bare = re.sub(r"^#thkr#", "", name_lower)
        for abap_file in _iter_all_abap_files(ABAP_DIR):
            if bare in abap_file.name.lower():
                matches.append(abap_file)

    if not matches:
        # List a sample of files to help Claude suggest the right name
        all_names = [f.name for f in sorted(_iter_all_abap_files(ABAP_DIR), key=lambda p: p.name)[:40]]
        return (
            f"File not found matching '{filename}'.\n"
            f"Use search_abap_code to locate the right file first.\n"
            f"Sample file names: {', '.join(all_names)}"
        )

    # Return the best (shortest name = most specific) match
    best = sorted(matches, key=lambda p: len(p.name))[0]
    rel = best.relative_to(ABAP_DIR)

    try:
        content = best.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"Error reading {rel}: {e}"

    # Cap at ~400 lines to avoid flooding context
    lines = content.splitlines()
    if len(lines) > 400:
        content = "\n".join(lines[:400]) + f"\n\n... ({len(lines) - 400} more lines truncated)"

    return f"File: {rel}\nLines: {len(lines)}\n\n{content}"


# ── Tool definitions (passed to Claude API) ───────────────────────────────────

TOOL_DEFINITIONS = [
    {
        "name": "search_interfaces",
        "description": (
            "Search AIF interface definitions by keyword. Searches across namespace (NS), "
            "interface name (IFNAME), and German description (IFDESC). "
            "Use to find which interfaces involve a specific external system or topic."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": (
                        "Search keyword — interface name fragment (e.g. '0004'), "
                        "external system name (e.g. 'EMSA', 'ELVIS'), "
                        "or German description keyword (e.g. 'Mahngericht', 'Polizei')"
                    ),
                }
            },
            "required": ["query"],
        },
    },
    {
        "name": "get_interface_details",
        "description": (
            "Retrieve full technical details for a specific interface: "
            "DDIC structure, check/init function modules, action pipeline (AIF_T_IFACT), "
            "function modules per action (AIF_T_FUNC), and field mappings (AIF_T_FMAP)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "ns": {
                    "type": "string",
                    "description": "Namespace, e.g. FREMDV or HAVWEB",
                },
                "ifname": {
                    "type": "string",
                    "description": "Interface name, e.g. I_0004_001 or O_0004_002",
                },
            },
            "required": ["ns", "ifname"],
        },
    },
    {
        "name": "get_interface_readme",
        "description": (
            "Retrieve pre-generated README documentation for an interface or domain. "
            "Faster than get_interface_details — check this first for documented interfaces. "
            "Try names like: I_0004_001, FREMDV, HAVWEB, EDAS, RUECKMELDUNG, SERiD."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "description": (
                        "Interface or domain name to look up. "
                        "E.g. 'I_0004_001', 'FREMDV', 'EDAS', 'O_0027_002'"
                    ),
                }
            },
            "required": ["name"],
        },
    },
    {
        "name": "search_abap_code",
        "description": (
            "Full-text search across ALL abapgit source files in the repository. "
            "Searches .abap (function modules, programs, classes, includes), "
            ".asddls (CDS views), .xslt.source.xml (XSLT transformations), "
            ".tabl.xml (DDIC structures/tables), .fugr.xml (function group metadata), "
            ".dtel.xml (data elements), .doma.xml (domains), and more. "
            "Use to find any object by name, field reference, or code pattern. "
            "To list CDS views: search_abap_code('cds_aif') or search_abap_code('define view')."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": (
                        "Text to search for — function module name, CDS view name, "
                        "XSLT name, structure name, field name, or any keyword. Case-insensitive."
                    ),
                },
                "context_lines": {
                    "type": "integer",
                    "description": "Lines of context around each match (default 15, max 30)",
                    "default": 15,
                },
            },
            "required": ["query"],
        },
    },
    {
        "name": "read_abap_file",
        "description": (
            "Read the complete source of any abapgit file matched by partial name. "
            "Works for ALL object types: .abap (FM/program/class), .asddls (CDS view), "
            ".xslt.source.xml (XSLT), .tabl.xml (structure), .fugr.xml (function group), "
            ".dtel.xml, .doma.xml, etc. "
            "Provide a partial filename — the tool does a case-insensitive partial match."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": (
                        "Partial filename to match, e.g. 'aif_fremdv_act_wrt_rk_er', "
                        "'cds_aif_ist_v2', 'laif_fremdv_maptop', or 'cl_aif_rueck'. "
                        "For CDS views omit the extension — just use the view name."
                    ),
                }
            },
            "required": ["filename"],
        },
    },
]


def dispatch_tool(name: str, inputs: dict) -> str:
    """Route a tool call by name to the matching implementation."""
    if name == "search_interfaces":
        return tool_search_interfaces(**inputs)
    if name == "get_interface_details":
        return tool_get_interface_details(**inputs)
    if name == "get_interface_readme":
        return tool_get_interface_readme(**inputs)
    if name == "search_abap_code":
        return tool_search_abap_code(**inputs)
    if name == "read_abap_file":
        return tool_read_abap_file(**inputs)
    return f"Unknown tool: {name}"
